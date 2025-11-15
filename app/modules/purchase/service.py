from app.modules.purchase import schemas as purchase_schemas
from fastapi import Depends, Request, HTTPException
from sqlalchemy.orm import Session
from app.core.database import get_db
from sqlalchemy import and_
from app.common import response as common_response
from typing import Union
from app.modules.purchase import models as purchase_models
from app.modules.setting import models as set_models
from sqlalchemy.orm import aliased
from app.modules.auth import models as auth_models
from app.modules.common import models as common_models
from app.modules.common import schemas as common_schemas
from app.common.schemas import request as common_request
from app.common.response import ApiResponse, PageResponse, ResponseBuilder
from sqlalchemy import func
from app.utils.auth_util import get_authenticated_user_no
from app.utils import com_code_util
from fastapi.responses import FileResponse
from app.utils.cj_logistics_util import request_cj_logistics_api
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import tempfile
import os
from datetime import datetime
from app.utils import alibaba_1688_util
from collections import defaultdict

def fetch_order_mst_list(
    filter: purchase_schemas.OrderMstFilterRequest,
    request: Request,
    pagination: common_request.PaginationRequest = Depends(),
    db: Session = Depends(get_db)
) -> ApiResponse[Union[PageResponse[purchase_schemas.OrderMstResponse], None]]:

    ComCode = common_models.ComCode
    ComCompany = auth_models.ComCompany

    # ComCode를 두 개의 별칭으로 생성
    ComCodePlatform = aliased(ComCode)  # platform_type용
    ComCodeStatus = aliased(ComCode)  # order_mst_status용

    query = (
        db.query(
            purchase_models.OrderMst,
            ComCodePlatform.code_name.label("platform_type_name"),
            ComCodeStatus.code_name.label("order_mst_status_name"),
            ComCompany.company_name.label("company_name")
        ).join(
            ComCodePlatform,
            purchase_models.OrderMst.platform_type_cd == ComCodePlatform.com_code
        ).join(
            ComCodeStatus,
            purchase_models.OrderMst.order_mst_status_cd == ComCodeStatus.com_code
        ).join(
            ComCompany,
            purchase_models.OrderMst.company_no == ComCompany.company_no
        ).filter(
            ComCodePlatform.parent_com_code == 'PLATFORM_TYPE_CD',
            ComCodePlatform.use_yn == 1,
            ComCodePlatform.del_yn == 0,
            ComCodeStatus.parent_com_code == 'ORDER_MST_STATUS_CD',
            ComCodeStatus.use_yn == 1,
            ComCodeStatus.del_yn == 0,
            purchase_models.OrderMst.del_yn == 0
        )
    )

    # 필터 조건 수정 - 빈 문자열 체크 추가
    if filter.order_memo and filter.order_memo.strip():
        query = query.filter(purchase_models.OrderMst.order_memo.like(f"%{filter.order_memo.strip()}%"))

    if filter.order_mst_status_cd and filter.order_mst_status_cd.strip():
        query = query.filter(purchase_models.OrderMst.order_mst_status_cd == filter.order_mst_status_cd.strip())

    # 날짜 필터링 로직 수정
    if filter.order_date_start and filter.order_date_start.strip():
        start_date = filter.order_date_start.strip()
        if filter.order_date_end and filter.order_date_end.strip():
            # 시작일과 종료일이 모두 있는 경우
            end_date = filter.order_date_end.strip()
            query = query.filter(
                purchase_models.OrderMst.order_date.between(start_date, end_date)
            )
        else:
            # 시작일만 있는 경우 - 시작일 이후 데이터
            query = query.filter(
                purchase_models.OrderMst.order_date >= start_date
            )
    elif filter.order_date_end and filter.order_date_end.strip():
        # 종료일만 있는 경우 - 종료일 이전 데이터
        end_date = filter.order_date_end.strip()
        query = query.filter(
            purchase_models.OrderMst.order_date <= end_date
        )

    query = query.order_by(purchase_models.OrderMst.updated_at.desc())

    # 전체 개수
    total_elements = query.count()

    # 페이징
    offset = (pagination.page - 1) * pagination.size
    orders = query.offset(offset).limit(pagination.size).all()

    order_list = []

    for order, platform_type_name, order_mst_status_name, company_name in orders:
        order_data = purchase_schemas.OrderMstResponse.from_orm(order)
        order_data.platform_type_name = platform_type_name
        order_data.order_mst_status_name = order_mst_status_name
        order_data.company_name = company_name
        order_list.append(order_data)

    return ResponseBuilder.paged_success(
        content=order_list,
        page=pagination.page,
        size=pagination.size,
        total_elements=total_elements
    )

def fetch_purchase_shipment_mst(
        request: Request,
        order_mst_no: int,
        db: Session = Depends(get_db)
) -> common_response.ApiResponse[Union[dict, None]]:
    try:

        shipment_mst_list = db.query(
            purchase_models.OrderShipmentMst
        ).filter(
            purchase_models.OrderShipmentMst.order_mst_no == order_mst_no,
            purchase_models.OrderShipmentMst.del_yn == 0
        ).order_by(
            purchase_models.OrderShipmentMst.estimated_yn.desc()
        ).all()
        # tabs 리스트를 for 루프 밖에서 초기화
        tabs = []

        for shipment_mst in shipment_mst_list:
            # 플랫폼 타입에 따른 구분
            if shipment_mst.platform_type_cd == "GROWTH":
                sub_label = shipment_mst.inbound_no
            else:
                sub_label = shipment_mst.edd

            tab = {
                "order_shipment_mst_no": shipment_mst.order_shipment_mst_no,
                "order_mst_no": shipment_mst.order_mst_no,
                "center_no": shipment_mst.center_no,
                "label": f"{shipment_mst.display_center_name}({sub_label})",
                "estimated_yn": shipment_mst.estimated_yn,
                "status": shipment_mst.order_shipment_mst_status_cd
            }
            tabs.append(tab)

        # data 변수를 for 루프 밖에서 정의
        data = {
            "tabs": tabs,
            "total_count": len(tabs),
            "order_mst_no": order_mst_no
        }

        return common_response.ResponseBuilder.success(
            data=data,
            message=f"쉽먼트 목록이 정상적으로 출력되었습니다. (총 {len(tabs)}건)"
        )

    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"쉽먼트 목록 출력 중 오류가 발생했습니다: {str(e)}"
        )


def fetch_shipment_dtl_list(
        order_shipment_mst_no: Union[str, int],
        request: Request,
        pagination: common_request.PaginationRequest,
        db: Session
) -> common_response.ApiResponse[Union[PageResponse[dict], None]]:
    """쉽먼트 마스터 번호로 특정 쉽먼트 DTL 조회 (MST, PACKING_DTL 정보 포함)"""
    try:
        # 쉽먼트 마스터 존재 확인
        existing_shipment_mst = db.query(purchase_models.OrderShipmentMst).filter(
            purchase_models.OrderShipmentMst.order_shipment_mst_no == order_shipment_mst_no,
            purchase_models.OrderShipmentMst.del_yn == 0
        ).first()

        if not existing_shipment_mst:
            raise HTTPException(
                status_code=400,
                detail="해당 쉽먼트를 찾을 수 없습니다.",
            )

        # center_name 서브쿼리
        center_subquery = db.query(set_models.SetCenter.center_name).filter(
            set_models.SetCenter.center_no == purchase_models.OrderShipmentMst.center_no,
            set_models.SetCenter.del_yn == 0
        ).scalar_subquery()

        # MST, DTL, PACKING_DTL, PACKING_MST LEFT OUTER JOIN 쿼리 구성
        query = db.query(
            purchase_models.OrderShipmentMst,
            purchase_models.OrderShipmentDtl,
            purchase_models.OrderShipmentPackingDtl,
            purchase_models.OrderShipmentPackingMst,
            center_subquery.label("center_name")
        ).join(
            purchase_models.OrderShipmentDtl,
            purchase_models.OrderShipmentMst.order_shipment_mst_no == purchase_models.OrderShipmentDtl.order_shipment_mst_no
        ).outerjoin(
            purchase_models.OrderShipmentPackingDtl,
            and_(
                purchase_models.OrderShipmentDtl.order_shipment_dtl_no == purchase_models.OrderShipmentPackingDtl.order_shipment_dtl_no,
                purchase_models.OrderShipmentPackingDtl.del_yn == 0  # JOIN 조건에 del_yn 포함
            )
        ).outerjoin(
            purchase_models.OrderShipmentPackingMst,
            and_(
                purchase_models.OrderShipmentPackingDtl.order_shipment_packing_mst_no == purchase_models.OrderShipmentPackingMst.order_shipment_packing_mst_no,
                purchase_models.OrderShipmentPackingMst.del_yn == 0  # JOIN 조건에 del_yn 포함
            )
        ).filter(
            purchase_models.OrderShipmentMst.order_shipment_mst_no == order_shipment_mst_no,
            purchase_models.OrderShipmentMst.del_yn == 0,
            purchase_models.OrderShipmentDtl.del_yn == 0
            # PACKING_DTL 필터 조건은 JOIN 조건으로 이동하여 LEFT JOIN이 제대로 동작하도록 함
        ).order_by(
            purchase_models.OrderShipmentDtl.created_at.desc(),
            purchase_models.OrderShipmentPackingDtl.created_at.desc()
        )

        # 전체 개수
        total_elements = query.count()

        # 페이징
        offset = (pagination.page - 1) * pagination.size
        results = query.offset(offset).limit(pagination.size).all()

        # 결과 데이터 변환
        dtl_data_list = []
        for mst, dtl, packing_dtl, packing_mst, center_name in results:
            combined_data = {
                # MST 정보
                "order_shipment_mst_no": mst.order_shipment_mst_no,
                "order_mst_no": mst.order_mst_no,
                "center_no": mst.center_no,
                "center_name": center_name,
                "edd": mst.edd,
                "order_shipment_mst_status_cd": mst.order_shipment_mst_status_cd,
                "mst_created_at": mst.created_at,
                "mst_created_by": mst.created_by,
                "mst_updated_at": mst.updated_at,
                "mst_updated_by": mst.updated_by,

                # DTL 정보
                "order_shipment_dtl_no": dtl.order_shipment_dtl_no,
                "order_shipment_packing_mst_no": dtl.order_shipment_packing_mst_no,
                "company_no": dtl.company_no,
                "order_number": dtl.order_number,
                "transport_type": dtl.transport_type,
                "sku_id": dtl.sku_id,
                "sku_barcode": dtl.sku_barcode,
                "sku_name": dtl.sku_name,
                "confirmed_quantity": dtl.confirmed_quantity,
                "shipped_quantity": dtl.shipped_quantity,
                "link": dtl.link,
                "option_type": dtl.option_type,
                "option_value": dtl.option_value,
                "linked_option": dtl.linked_option,
                "linked_spec_id": dtl.linked_spec_id,
                "linked_sku_id": dtl.linked_sku_id,
                "linked_open_uid": dtl.linked_open_uid,
                "multiple_value": dtl.multiple_value,
                "length_mm": float(dtl.length_mm) if dtl.length_mm else None,
                "width_mm": float(dtl.width_mm) if dtl.width_mm else None,
                "height_mm": float(dtl.height_mm) if dtl.height_mm else None,
                "weight_g": float(dtl.weight_g) if dtl.weight_g else None,
                "inspected_quantity": dtl.inspected_quantity,
                "packing_tracking_number": dtl.purchase_tracking_number,
                "virtual_packed_yn": dtl.virtual_packed_yn,
                "del_yn": dtl.del_yn,

                # PACKING_DTL 정보 (LEFT JOIN으로 가져온 값들)
                "order_shipment_packing_dtl_no": packing_dtl.order_shipment_packing_dtl_no if packing_dtl else None,
                "packing_quantity": packing_dtl.packing_quantity if packing_dtl else None,
                "tracking_number": packing_dtl.tracking_number if packing_dtl else None,

                # PACKING_MST 정보 (박스 정보)
                "box_name": packing_mst.box_name if packing_mst else None,
                "package_box_spec_cd": packing_mst.package_box_spec_cd if packing_mst else None,

                # PACKING_DTL 생성/수정 정보
                "packing_dtl_created_at": packing_dtl.created_at if packing_dtl else None,
                "packing_dtl_created_by": packing_dtl.created_by if packing_dtl else None,
                "packing_dtl_updated_at": packing_dtl.updated_at if packing_dtl else None,
                "packing_dtl_updated_by": packing_dtl.updated_by if packing_dtl else None
            }
            dtl_data_list.append(combined_data)

        return ResponseBuilder.paged_success(
            content=dtl_data_list,
            page=pagination.page,
            size=pagination.size,
            total_elements=total_elements
        )

    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"쉽먼트 DTL 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )


def fetch_shipment_estimate_product_list(
        order_shipment_mst_no: Union[str, int],
        request: Request,
        pagination: common_request.PaginationRequest,
        db: Session
) -> common_response.ApiResponse[Union[PageResponse[dict], None]]:
    """쉽먼트 마스터 번호로 견적 상품 정보 조회 (estimated_yn이 1일 때)"""
    try:
        # 쉽먼트 마스터 존재 확인
        existing_shipment_mst = db.query(purchase_models.OrderShipmentMst).filter(
            purchase_models.OrderShipmentMst.order_shipment_mst_no == order_shipment_mst_no,
            purchase_models.OrderShipmentMst.del_yn == 0
        ).first()

        if not existing_shipment_mst:
            raise HTTPException(
                status_code=400,
                detail="해당 쉽먼트를 찾을 수 없습니다.",
            )

        # estimated_yn 확인
        if existing_shipment_mst.estimated_yn != 1:
            raise HTTPException(
                status_code=400,
                detail="견적이 생성되지 않은 쉽먼트입니다.",
            )

        # center_name 서브쿼리
        center_subquery = db.query(set_models.SetCenter.center_name).filter(
            set_models.SetCenter.center_no == purchase_models.OrderShipmentMst.center_no,
            set_models.SetCenter.del_yn == 0
        ).scalar_subquery()

        # 필요한 컬럼만 명시적으로 선택 (중복 컬럼은 label로 구분)
        query = (db.query(
            # EstimateProduct 컬럼
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_product_no,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_no,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_mst_no,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no,
            purchase_models.OrderShipmentEstimateProduct.company_no,
            purchase_models.OrderShipmentEstimateProduct.center_no,
            purchase_models.OrderShipmentEstimateProduct.sku_id,
            purchase_models.OrderShipmentEstimateProduct.sku_name,
            purchase_models.OrderShipmentEstimateProduct.bundle,
            purchase_models.OrderShipmentEstimateProduct.purchase_quantity,
            purchase_models.OrderShipmentEstimateProduct.product_unit_price,
            purchase_models.OrderShipmentEstimateProduct.product_total_amount.label("product_product_total_amount"),
            purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_cd,
            purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_unit_price,
            purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_total_amount,
            purchase_models.OrderShipmentEstimateProduct.fail_yn,
            purchase_models.OrderShipmentEstimateProduct.total_amount.label("product_total_amount"),
            purchase_models.OrderShipmentEstimateProduct.remark,
            purchase_models.OrderShipmentEstimateProduct.platform_type_cd.label("product_platform_type_cd"),
            purchase_models.OrderShipmentEstimateProduct.created_at.label("product_created_at"),
            purchase_models.OrderShipmentEstimateProduct.created_by.label("product_created_by"),
            purchase_models.OrderShipmentEstimateProduct.updated_at.label("product_updated_at"),
            purchase_models.OrderShipmentEstimateProduct.updated_by.label("product_updated_by"),

            # Estimate 컬럼
            purchase_models.OrderShipmentEstimate.estimate_id,
            purchase_models.OrderShipmentEstimate.estimate_date,
            purchase_models.OrderShipmentEstimate.product_total_amount.label("estimate_product_total_amount"),
            purchase_models.OrderShipmentEstimate.vinyl_total_amount,
            purchase_models.OrderShipmentEstimate.box_total_amount,
            purchase_models.OrderShipmentEstimate.estimate_total_amount,

            # ShipmentMst 컬럼
            purchase_models.OrderShipmentMst.inbound_id,
            purchase_models.OrderShipmentMst.inbound_no,
            purchase_models.OrderShipmentMst.display_center_name,
            purchase_models.OrderShipmentMst.edd,
            purchase_models.OrderShipmentMst.order_shipment_mst_status_cd,
            purchase_models.OrderShipmentMst.estimated_yn,
            center_subquery.label("center_name"),

            # ShipmentDtl 컬럼 (선택적)
            purchase_models.OrderShipmentDtl.order_number,
            purchase_models.OrderShipmentDtl.sku_barcode,
            purchase_models.OrderShipmentDtl.confirmed_quantity,
            purchase_models.OrderShipmentDtl.shipped_quantity,
            purchase_models.OrderShipmentDtl.link,
            purchase_models.OrderShipmentDtl.option_type,
            purchase_models.OrderShipmentDtl.option_value,
            purchase_models.OrderShipmentDtl.length_mm,
            purchase_models.OrderShipmentDtl.width_mm,
            purchase_models.OrderShipmentDtl.height_mm,
            purchase_models.OrderShipmentDtl.weight_g,
            purchase_models.OrderShipmentDtl.coupang_option_name,
            purchase_models.OrderShipmentDtl.coupang_product_id,
            purchase_models.OrderShipmentDtl.coupang_option_id,
            purchase_models.OrderShipmentDtl.transport_type,
            purchase_models.OrderShipmentDtl.purchase_tracking_number,
            purchase_models.OrderShipmentDtl.purchase_order_number,

            # PackingDtl 컬럼
            purchase_models.OrderShipmentPackingDtl.box_name,
            purchase_models.OrderShipmentPackingDtl.packing_quantity,
            purchase_models.OrderShipmentPackingDtl.tracking_number,

            # ✅ PackingMst 컬럼 추가
            purchase_models.OrderShipmentPackingMst.order_shipment_packing_mst_no
        ).join(
            purchase_models.OrderShipmentEstimate,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_no == purchase_models.OrderShipmentEstimate.order_shipment_estimate_no
        ).join(
            purchase_models.OrderShipmentMst,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_mst_no == purchase_models.OrderShipmentMst.order_shipment_mst_no
        ).outerjoin(
            purchase_models.OrderShipmentDtl,
            and_(
                purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no == purchase_models.OrderShipmentDtl.order_shipment_dtl_no,
                purchase_models.OrderShipmentDtl.del_yn == 0
            )
        ).outerjoin(
            purchase_models.OrderShipmentPackingDtl,
            and_(
                purchase_models.OrderShipmentDtl.order_shipment_dtl_no == purchase_models.OrderShipmentPackingDtl.order_shipment_dtl_no,
                purchase_models.OrderShipmentPackingDtl.del_yn == 0
            )
        ).outerjoin(  # ✅ PackingMst 조인 추가
            purchase_models.OrderShipmentPackingMst,
            and_(
                purchase_models.OrderShipmentPackingDtl.order_shipment_packing_mst_no == purchase_models.OrderShipmentPackingMst.order_shipment_packing_mst_no,
                purchase_models.OrderShipmentPackingMst.del_yn == 0
            )
        ).filter(
            purchase_models.OrderShipmentEstimateProduct.order_shipment_mst_no == order_shipment_mst_no,
            purchase_models.OrderShipmentEstimateProduct.del_yn == 0,
            purchase_models.OrderShipmentEstimate.del_yn == 0,
            purchase_models.OrderShipmentMst.del_yn == 0
        ).order_by(
            purchase_models.OrderShipmentEstimateProduct.created_at.desc()
        ))

        # 전체 개수
        total_elements = query.count()

        # 페이징
        offset = (pagination.page - 1) * pagination.size
        results = query.offset(offset).limit(pagination.size).all()

        # 결과 데이터 변환
        estimate_product_list = []
        for row in results:
            combined_data = {
                # 견적 상품 정보
                "order_shipment_estimate_product_no": row.order_shipment_estimate_product_no,
                "order_shipment_estimate_no": row.order_shipment_estimate_no,
                "order_shipment_mst_no": row.order_shipment_mst_no,
                "order_shipment_dtl_no": row.order_shipment_dtl_no,
                "company_no": row.company_no,
                "center_no": row.center_no,
                "center_name": row.center_name,
                "sku_id": row.sku_id,
                "sku_name": row.sku_name,
                "bundle": row.bundle,
                "purchase_quantity": row.purchase_quantity,
                "product_unit_price": float(row.product_unit_price) if row.product_unit_price else 0.0,
                "product_product_total_amount": float(
                    row.product_product_total_amount) if row.product_product_total_amount else 0.0,
                "package_vinyl_spec_cd": row.package_vinyl_spec_cd,
                "package_vinyl_spec_unit_price": float(
                    row.package_vinyl_spec_unit_price) if row.package_vinyl_spec_unit_price else 0.0,
                "package_vinyl_spec_total_amount": float(
                    row.package_vinyl_spec_total_amount) if row.package_vinyl_spec_total_amount else 0.0,
                "fail_yn": row.fail_yn,
                "total_amount": float(row.product_total_amount) if row.product_total_amount else 0.0,
                "remark": row.remark,
                "platform_type_cd": row.product_platform_type_cd,

                # 견적서 정보
                "estimate_id": row.estimate_id,
                "estimate_date": row.estimate_date,
                "estimate_total_amount": float(row.estimate_total_amount) if row.estimate_total_amount else 0.0,
                "estimate_product_total_amount": float(
                    row.estimate_product_total_amount) if row.estimate_product_total_amount else 0.0,
                "vinyl_total_amount": float(row.vinyl_total_amount) if row.vinyl_total_amount else 0.0,
                "box_total_amount": float(row.box_total_amount) if row.box_total_amount else 0.0,

                # 쉽먼트 MST 정보
                "inbound_id": row.inbound_id,
                "inbound_no": row.inbound_no,
                "display_center_name": row.display_center_name,
                "edd": row.edd,
                "order_shipment_mst_status_cd": row.order_shipment_mst_status_cd,
                "estimated_yn": row.estimated_yn,

                # 쉽먼트 DTL 정보
                "order_number": row.order_number,
                "sku_barcode": row.sku_barcode if row.sku_barcode else None,
                "confirmed_quantity": row.confirmed_quantity if row.confirmed_quantity else None,
                "shipped_quantity": row.shipped_quantity if row.shipped_quantity else None,
                "link": row.link if row.link else None,
                "option_type": row.option_type if row.option_type else None,
                "option_value": row.option_value if row.option_value else None,
                "length_mm": float(row.length_mm) if row.length_mm else None,
                "width_mm": float(row.width_mm) if row.width_mm else None,
                "height_mm": float(row.height_mm) if row.height_mm else None,
                "weight_g": float(row.weight_g) if row.weight_g else None,
                "coupang_option_name": row.coupang_option_name if row.coupang_option_name else None,
                "coupang_product_id": row.coupang_product_id if row.coupang_product_id else None,
                "coupang_option_id": row.coupang_option_id if row.coupang_option_id else None,
                "transport_type": row.transport_type if row.transport_type else None,
                "packing_quantity": row.packing_quantity if row.packing_quantity else None,
                "purchase_tracking_number": row.purchase_tracking_number if row.purchase_tracking_number else None,
                "tracking_number": row.tracking_number if row.tracking_number else None,
                "purchase_order_number": row.purchase_order_number if row.purchase_order_number else None,

                # Packing 정보
                "box_name": row.box_name if row.box_name else None,
                "order_shipment_packing_mst_no": row.order_shipment_packing_mst_no if row.order_shipment_packing_mst_no else None,

                # 생성/수정 정보
                "created_at": row.product_created_at,
                "created_by": row.product_created_by,
                "updated_at": row.product_updated_at,
                "updated_by": row.product_updated_by
            }
            estimate_product_list.append(combined_data)

        return ResponseBuilder.paged_success(
            content=estimate_product_list,
            page=pagination.page,
            size=pagination.size,
            total_elements=total_elements
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"견적 상품 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )
    

def fetch_shipment_dtl_all_list(
    order_mst_no: Union[str, int],
    request: Request,
    pagination: common_request.PaginationRequest,
    db: Session
) -> common_response.ApiResponse[Union[PageResponse[dict], None]]:
    """발주서 마스터 번호로 모든 쉽먼트 DTL 조회 (MST, PACKING_DTL 정보 포함)"""
    try:
        # 발주서 마스터 존재 확인
        existing_growth_order_mst = db.query(purchase_models.OrderMst).filter(
            purchase_models.OrderMst.order_mst_no == order_mst_no,
            purchase_models.OrderMst.del_yn == 0
        ).first()

        if not existing_growth_order_mst:
            raise HTTPException(
                status_code=400,
                detail="해당 발주서를 찾을 수 없습니다."
            )

        # center_name 서브쿼리
        center_subquery = db.query(set_models.SetCenter.center_name).filter(
            set_models.SetCenter.center_no == purchase_models.OrderShipmentMst.center_no,
            set_models.SetCenter.del_yn == 0
        ).scalar_subquery()

        # MST, DTL, PACKING_DTL, PACKING_MST LEFT OUTER JOIN 쿼리 구성
        query = db.query(
            purchase_models.OrderShipmentMst,
            purchase_models.OrderShipmentDtl,
            purchase_models.OrderShipmentPackingDtl,
            purchase_models.OrderShipmentPackingMst,
            center_subquery.label("center_name")
        ).join(
            purchase_models.OrderShipmentDtl,
            purchase_models.OrderShipmentMst.order_shipment_mst_no == purchase_models.OrderShipmentDtl.order_shipment_mst_no
        ).outerjoin(
            purchase_models.OrderShipmentPackingDtl,
            and_(
                purchase_models.OrderShipmentDtl.order_shipment_dtl_no == purchase_models.OrderShipmentPackingDtl.order_shipment_dtl_no,
                purchase_models.OrderShipmentPackingDtl.del_yn == 0  # JOIN 조건에 del_yn 포함
            )
        ).outerjoin(
            purchase_models.OrderShipmentPackingMst,
            and_(
                purchase_models.OrderShipmentPackingDtl.order_shipment_packing_mst_no == purchase_models.OrderShipmentPackingMst.order_shipment_packing_mst_no,
                purchase_models.OrderShipmentPackingMst.del_yn == 0  # JOIN 조건에 del_yn 포함
            )
        ).filter(
            purchase_models.OrderShipmentMst.order_mst_no == order_mst_no,
            purchase_models.OrderShipmentMst.del_yn == 0,
            purchase_models.OrderShipmentDtl.del_yn == 0
            # PACKING_DTL 필터 조건은 JOIN 조건으로 이동하여 LEFT JOIN이 제대로 동작하도록 함
        ).order_by(
            purchase_models.OrderShipmentMst.estimated_yn.desc(),
            purchase_models.OrderShipmentDtl.created_at.desc(),
            purchase_models.OrderShipmentPackingDtl.created_at.desc()
        )

        # 전체 개수
        total_elements = query.count()

        # 페이징
        offset = (pagination.page - 1) * pagination.size
        results = query.offset(offset).limit(pagination.size).all()

        # 공통코드
        shipment_status_com_code_dict = com_code_util.get_com_code_dict_by_parent_code("ORDER_SHIPMENT_MST_STATUS_CD", db)

        # 결과 데이터 변환
        dtl_data_list = []
        for mst, dtl, packing_dtl, packing_mst, center_name in results:
            com_code = shipment_status_com_code_dict.get(mst.order_shipment_mst_status_cd)
            combined_data = {
                # MST 정보
                "order_shipment_mst_no": mst.order_shipment_mst_no,
                "order_mst_no": mst.order_mst_no,
                "center_no": mst.center_no,
                "estimated_yn": mst.estimated_yn,
                "center_name": center_name,
                "edd": mst.edd,
                "order_shipment_mst_status_cd": mst.order_shipment_mst_status_cd,
                "order_shipment_mst_status_name": com_code.code_name,
                "order_shipment_mst_status_color": com_code.keyword1,
                "mst_created_at": mst.created_at,
                "mst_created_by": mst.created_by,
                "mst_updated_at": mst.updated_at,
                "mst_updated_by": mst.updated_by,

                # DTL 정보
                "order_shipment_dtl_no": dtl.order_shipment_dtl_no,
                "order_shipment_packing_mst_no": dtl.order_shipment_packing_mst_no,
                "company_no": dtl.company_no,
                "order_number": dtl.order_number,
                "transport_type": dtl.transport_type,
                "sku_id": dtl.sku_id,
                "sku_barcode": dtl.sku_barcode,
                "sku_name": dtl.sku_name,
                "confirmed_quantity": dtl.confirmed_quantity,
                "purchase_tracking_number": dtl.purchase_tracking_number,  # SHIPMENT_DTL의 1688 운송장번호
                "shipped_quantity": dtl.shipped_quantity,
                "link": dtl.link,
                "option_type": dtl.option_type,
                "option_value": dtl.option_value,
                "linked_option": dtl.linked_option,
                "linked_spec_id": dtl.linked_spec_id,
                "linked_sku_id": dtl.linked_sku_id,
                "linked_open_uid": dtl.linked_open_uid,
                "multiple_value": dtl.multiple_value,
                "length_mm": float(dtl.length_mm) if dtl.length_mm else None,
                "width_mm": float(dtl.width_mm) if dtl.width_mm else None,
                "height_mm": float(dtl.height_mm) if dtl.height_mm else None,
                "weight_g": float(dtl.weight_g) if dtl.weight_g else None,
                "inspected_quantity": dtl.inspected_quantity,
                "virtual_packed_yn": dtl.virtual_packed_yn,
                "dtl_created_at": dtl.created_at,
                "dtl_created_by": dtl.created_by,
                "dtl_updated_at": dtl.updated_at,
                "dtl_updated_by": dtl.updated_by,

                # PACKING_DTL 정보 (LEFT JOIN으로 가져온 값들)
                "order_shipment_packing_dtl_no": packing_dtl.order_shipment_packing_dtl_no if packing_dtl else None,
                "packing_quantity": packing_dtl.packing_quantity if packing_dtl else None,
                "packing_tracking_number": packing_dtl.tracking_number if packing_dtl else None,  # PACKING_DTL의 tracking_number

                # PACKING_MST 정보 (박스 정보)
                "box_name": packing_mst.box_name if packing_mst else None,
                "package_box_spec_cd": packing_mst.package_box_spec_cd if packing_mst else None,

                # PACKING_DTL 생성/수정 정보
                "tracking_number": packing_dtl.tracking_number if packing_dtl else None,
                "packing_dtl_created_at": packing_dtl.created_at if packing_dtl else None,
                "packing_dtl_created_by": packing_dtl.created_by if packing_dtl else None,
                "packing_dtl_updated_at": packing_dtl.updated_at if packing_dtl else None,
                "packing_dtl_updated_by": packing_dtl.updated_by if packing_dtl else None
            }
            dtl_data_list.append(combined_data)

        return ResponseBuilder.paged_success(
            content=dtl_data_list,
            page=pagination.page,
            size=pagination.size,
            total_elements=total_elements
        )

    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"쉽먼트 DTL 전체 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )


def fetch_shipment_estimate_product_list_all(
        order_mst_no: Union[str, int],
        request: Request,
        pagination: common_request.PaginationRequest,
        db: Session
) -> common_response.ApiResponse[Union[PageResponse[dict], None]]:
    """발주서 마스터 번호로 모든 견적 상품 정보 조회 (estimated_yn이 1인 모든 shipment의 견적 데이터)"""
    try:
        # 발주서 마스터 존재 확인
        existing_order_mst = db.query(purchase_models.OrderMst).filter(
            purchase_models.OrderMst.order_mst_no == order_mst_no,
            purchase_models.OrderMst.del_yn == 0
        ).first()

        if not existing_order_mst:
            raise HTTPException(
                status_code=400,
                detail="해당 발주서를 찾을 수 없습니다.",
            )

        # center_name 서브쿼리
        center_subquery = db.query(set_models.SetCenter.center_name).filter(
            set_models.SetCenter.center_no == purchase_models.OrderShipmentMst.center_no,
            set_models.SetCenter.del_yn == 0
        ).scalar_subquery()

        # 쉽먼트 상태명 서브쿼리
        shipment_status_subquery = db.query(common_models.ComCode.code_name).filter(
            common_models.ComCode.com_code == purchase_models.OrderShipmentMst.order_shipment_mst_status_cd,
            common_models.ComCode.parent_com_code == 'ORDER_SHIPMENT_MST_STATUS_CD',
            common_models.ComCode.del_yn == 0,
            common_models.ComCode.use_yn == 1
        ).scalar_subquery()

        # 포장비닐 사양명 서브쿼리
        vinyl_spec_subquery = db.query(common_models.ComCode.code_name).filter(
            common_models.ComCode.com_code == purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_cd,
            common_models.ComCode.parent_com_code == 'PACKAGE_VINYL_SPEC_CD',
            common_models.ComCode.del_yn == 0,
            common_models.ComCode.use_yn == 1
        ).scalar_subquery()

        # 필요한 컬럼만 명시적으로 선택 (중복 컬럼은 label로 구분)
        query = (db.query(
            # EstimateProduct 컬럼
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_product_no,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_no,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_mst_no,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no,
            purchase_models.OrderShipmentEstimateProduct.company_no,
            purchase_models.OrderShipmentEstimateProduct.center_no,
            purchase_models.OrderShipmentEstimateProduct.sku_id,
            purchase_models.OrderShipmentEstimateProduct.sku_name,
            purchase_models.OrderShipmentEstimateProduct.bundle,
            purchase_models.OrderShipmentEstimateProduct.purchase_quantity,
            purchase_models.OrderShipmentEstimateProduct.product_unit_price,
            purchase_models.OrderShipmentEstimateProduct.product_total_amount.label("product_product_total_amount"),
            purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_cd,
            purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_unit_price,
            purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_total_amount,
            purchase_models.OrderShipmentEstimateProduct.fail_yn,
            purchase_models.OrderShipmentEstimateProduct.total_amount.label("product_total_amount"),
            purchase_models.OrderShipmentEstimateProduct.remark,
            purchase_models.OrderShipmentEstimateProduct.platform_type_cd.label("product_platform_type_cd"),
            purchase_models.OrderShipmentEstimateProduct.created_at.label("product_created_at"),
            purchase_models.OrderShipmentEstimateProduct.created_by.label("product_created_by"),
            purchase_models.OrderShipmentEstimateProduct.updated_at.label("product_updated_at"),
            purchase_models.OrderShipmentEstimateProduct.updated_by.label("product_updated_by"),

            # Estimate 컬럼
            purchase_models.OrderShipmentEstimate.order_mst_no,
            purchase_models.OrderShipmentEstimate.estimate_id,
            purchase_models.OrderShipmentEstimate.estimate_date,
            purchase_models.OrderShipmentEstimate.product_total_amount.label("estimate_product_total_amount"),
            purchase_models.OrderShipmentEstimate.vinyl_total_amount,
            purchase_models.OrderShipmentEstimate.box_total_amount,
            purchase_models.OrderShipmentEstimate.estimate_total_amount,

            # ShipmentMst 컬럼
            purchase_models.OrderShipmentMst.inbound_id,
            purchase_models.OrderShipmentMst.inbound_no,
            purchase_models.OrderShipmentMst.display_center_name,
            purchase_models.OrderShipmentMst.edd,
            purchase_models.OrderShipmentMst.order_shipment_mst_status_cd,
            purchase_models.OrderShipmentMst.estimated_yn,
            center_subquery.label("center_name"),
            shipment_status_subquery.label("order_shipment_mst_status_name"),
            vinyl_spec_subquery.label("package_vinyl_spec_name"),

            # ShipmentDtl 컬럼 (선택적)
            purchase_models.OrderShipmentDtl.order_number,
            purchase_models.OrderShipmentDtl.sku_barcode,
            purchase_models.OrderShipmentDtl.confirmed_quantity.label("dtl_confirmed_quantity"),
            purchase_models.OrderShipmentDtl.shipped_quantity,
            purchase_models.OrderShipmentDtl.link,
            purchase_models.OrderShipmentDtl.option_type,
            purchase_models.OrderShipmentDtl.option_value,
            purchase_models.OrderShipmentDtl.length_mm,
            purchase_models.OrderShipmentDtl.width_mm,
            purchase_models.OrderShipmentDtl.height_mm,
            purchase_models.OrderShipmentDtl.weight_g,
            purchase_models.OrderShipmentDtl.coupang_option_name,
            purchase_models.OrderShipmentDtl.coupang_product_id,
            purchase_models.OrderShipmentDtl.coupang_option_id,
            purchase_models.OrderShipmentDtl.transport_type,
            purchase_models.OrderShipmentDtl.linked_open_uid,
            purchase_models.OrderShipmentDtl.purchase_tracking_number,
            purchase_models.OrderShipmentDtl.purchase_order_number,

            # PackingDtl 컬럼
            purchase_models.OrderShipmentPackingDtl.packing_quantity,
            purchase_models.OrderShipmentPackingDtl.box_name,
            purchase_models.OrderShipmentPackingDtl.tracking_number,

            # PackingMst 컬럼
            purchase_models.OrderShipmentPackingMst.order_shipment_packing_mst_no,
        ).join(
            purchase_models.OrderShipmentEstimate,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_no == purchase_models.OrderShipmentEstimate.order_shipment_estimate_no
        ).join(
            purchase_models.OrderShipmentMst,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_mst_no == purchase_models.OrderShipmentMst.order_shipment_mst_no
        ).outerjoin(
            purchase_models.OrderShipmentDtl,
            and_(
                purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no == purchase_models.OrderShipmentDtl.order_shipment_dtl_no,
                purchase_models.OrderShipmentDtl.del_yn == 0
            )
        ).outerjoin(
            purchase_models.OrderShipmentPackingDtl,
            and_(
                purchase_models.OrderShipmentDtl.order_shipment_dtl_no == purchase_models.OrderShipmentPackingDtl.order_shipment_dtl_no,
                purchase_models.OrderShipmentPackingDtl.del_yn == 0
            )
        ).outerjoin(  # ✅ PackingMst 조인 추가
            purchase_models.OrderShipmentPackingMst,
            and_(
                purchase_models.OrderShipmentPackingDtl.order_shipment_packing_mst_no == purchase_models.OrderShipmentPackingMst.order_shipment_packing_mst_no,
                purchase_models.OrderShipmentPackingMst.del_yn == 0
            )
        ).filter(
            purchase_models.OrderShipmentEstimate.order_mst_no == order_mst_no,
            purchase_models.OrderShipmentEstimateProduct.del_yn == 0,
            purchase_models.OrderShipmentEstimate.del_yn == 0,
            purchase_models.OrderShipmentMst.del_yn == 0
        ).order_by(
            purchase_models.OrderShipmentMst.estimated_yn.desc(),
            purchase_models.OrderShipmentEstimateProduct.created_at.desc()
        ))

        # 전체 개수
        total_elements = query.count()

        # 페이징
        offset = (pagination.page - 1) * pagination.size
        results = query.offset(offset).limit(pagination.size).all()

        # 결과 데이터 변환
        estimate_product_list = []
        for row in results:
            combined_data = {
                # 견적 상품 정보
                "order_shipment_estimate_product_no": row.order_shipment_estimate_product_no,
                "order_shipment_estimate_no": row.order_shipment_estimate_no,
                "order_shipment_mst_no": row.order_shipment_mst_no,
                "order_shipment_dtl_no": row.order_shipment_dtl_no,
                "company_no": row.company_no,
                "center_no": row.center_no,
                "center_name": row.center_name,
                "sku_id": row.sku_id,
                "sku_name": row.sku_name,
                "bundle": row.bundle,
                "purchase_quantity": row.purchase_quantity,
                "product_unit_price": float(row.product_unit_price) if row.product_unit_price else 0.0,
                "product_product_total_amount": float(
                    row.product_product_total_amount) if row.product_product_total_amount else 0.0,
                "package_vinyl_spec_cd": row.package_vinyl_spec_cd,
                "package_vinyl_spec_name": row.package_vinyl_spec_name,
                "package_vinyl_spec_unit_price": float(
                    row.package_vinyl_spec_unit_price) if row.package_vinyl_spec_unit_price else 0.0,
                "package_vinyl_spec_total_amount": float(
                    row.package_vinyl_spec_total_amount) if row.package_vinyl_spec_total_amount else 0.0,
                "fail_yn": row.fail_yn,
                "total_amount": float(row.product_total_amount) if row.product_total_amount else 0.0,
                "remark": row.remark,
                "platform_type_cd": row.product_platform_type_cd,

                # 견적서 정보
                "order_mst_no": row.order_mst_no,
                "estimate_id": row.estimate_id,
                "estimate_date": row.estimate_date,
                "estimate_total_amount": float(row.estimate_total_amount) if row.estimate_total_amount else 0.0,
                "estimate_product_total_amount": float(
                    row.estimate_product_total_amount) if row.estimate_product_total_amount else 0.0,
                "vinyl_total_amount": float(row.vinyl_total_amount) if row.vinyl_total_amount else 0.0,
                "box_total_amount": float(row.box_total_amount) if row.box_total_amount else 0.0,

                # 쉽먼트 MST 정보
                "inbound_id": row.inbound_id,
                "inbound_no": row.inbound_no,
                "display_center_name": row.display_center_name,
                "edd": row.edd,
                "order_shipment_mst_status_cd": row.order_shipment_mst_status_cd,
                "order_shipment_mst_status_name": row.order_shipment_mst_status_name,
                "estimated_yn": row.estimated_yn,

                # 쉽먼트 DTL 정보
                "order_number": row.order_number if row.order_number else None,
                "sku_barcode": row.sku_barcode if row.sku_barcode else None,
                "confirmed_quantity": row.dtl_confirmed_quantity if row.dtl_confirmed_quantity else None,
                "shipped_quantity": row.shipped_quantity if row.shipped_quantity else None,
                "link": row.link if row.link else None,
                "option_type": row.option_type if row.option_type else None,
                "option_value": row.option_value if row.option_value else None,
                "length_mm": float(row.length_mm) if row.length_mm else None,
                "width_mm": float(row.width_mm) if row.width_mm else None,
                "height_mm": float(row.height_mm) if row.height_mm else None,
                "weight_g": float(row.weight_g) if row.weight_g else None,
                "coupang_option_name": row.coupang_option_name if row.coupang_option_name else None,
                "coupang_product_id": row.coupang_product_id if row.coupang_product_id else None,
                "coupang_option_id": row.coupang_option_id if row.coupang_option_id else None,
                "transport_type": row.transport_type if row.transport_type else None,
                "linked_open_uid": row.linked_open_uid if row.linked_open_uid else None,
                "purchase_tracking_number": row.purchase_tracking_number if row.purchase_tracking_number else None,
                "purchase_order_number": row.purchase_order_number if row.purchase_order_number else None,

                # Packing 정보
                "packing_quantity": row.packing_quantity if row.packing_quantity else None,
                "box_name": row.box_name if row.box_name else None,
                "tracking_number": row.tracking_number if row.tracking_number else None,
                "order_shipment_packing_mst_no": row.order_shipment_packing_mst_no if row.order_shipment_packing_mst_no else None,

                # 생성/수정 정보
                "created_at": row.product_created_at,
                "created_by": row.product_created_by,
                "updated_at": row.product_updated_at,
                "updated_by": row.product_updated_by
            }
            estimate_product_list.append(combined_data)

        return ResponseBuilder.paged_success(
            content=estimate_product_list,
            page=pagination.page,
            size=pagination.size,
            total_elements=total_elements
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"견적 상품 전체 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )


def fetch_estimate_mst_list(
        order_mst_no: Union[str, int],
        pagination: common_request.PaginationRequest,
        request: Request,
        db: Session
) -> common_response.ApiResponse[Union[PageResponse[dict], None]]:
    """견적서 목록 조회"""
    try:

        # 1. 발주서 마스터 존재 확인
        rocket_order_mst = db.query(purchase_models.OrderMst).filter(
            purchase_models.OrderMst.order_mst_no == order_mst_no,
            purchase_models.OrderMst.del_yn == 0
        ).first()

        if not rocket_order_mst:
            raise HTTPException(
                status_code=404,
                detail="해당 발주서를 찾을 수 없습니다."
            )

        # 2. 견적서 목록 쿼리
        query = db.query(purchase_models.OrderShipmentEstimate).filter(
            purchase_models.OrderShipmentEstimate.order_mst_no == order_mst_no,
            purchase_models.OrderShipmentEstimate.del_yn == 0
        )

        # 정렬 (최신순)
        query = query.order_by(purchase_models.OrderShipmentEstimate.created_at.desc())

        # 전체 개수
        total_elements = query.count()

        # 페이징
        offset = (pagination.page - 1) * pagination.size
        estimates = query.offset(offset).limit(pagination.size).all()

        # 3. 데이터 포맷팅
        estimate_list = []
        for estimate in estimates:
            estimate_data = {
                "order_shipment_estimate_no": estimate.order_shipment_estimate_no,
                "order_mst_no": estimate.order_mst_no,
                "company_no": estimate.company_no,
                "estimate_id": estimate.estimate_id,
                "estimate_date": estimate.estimate_date,
                "product_total_amount": float(estimate.product_total_amount) if estimate.product_total_amount else 0.0,
                "vinyl_total_amount": float(estimate.vinyl_total_amount) if estimate.vinyl_total_amount else 0.0,
                "box_total_amount": float(estimate.box_total_amount) if estimate.box_total_amount else 0.0,
                "estimate_total_amount": float(estimate.estimate_total_amount) if estimate.estimate_total_amount else 0.0,
                "created_at": estimate.created_at.isoformat() if estimate.created_at else None,
                "created_by": estimate.created_by,
                "updated_at": estimate.updated_at.isoformat() if estimate.updated_at else None,
                "updated_by": estimate.updated_by
            }
            estimate_list.append(estimate_data)

        return common_response.ResponseBuilder.paged_success(
            content=estimate_list,
            page=pagination.page,
            size=pagination.size,
            total_elements=total_elements
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"견적서 목록 조회 중 오류가 발생했습니다: {str(e)}"
        )

def fetch_estimate_dtl(
        order_shipment_estimate_no: Union[str, int],
        request: Request,
        db: Session
) -> common_response.ApiResponse[dict]:
    """견적서 상세 조회"""
    try:
        # 1. 견적서 마스터 존재 확인
        estimate_mst = db.query(purchase_models.OrderShipmentEstimate).filter(
            purchase_models.OrderShipmentEstimate.order_shipment_estimate_no == order_shipment_estimate_no,
            purchase_models.OrderShipmentEstimate.del_yn == 0
        ).first()

        if not estimate_mst:
            raise HTTPException(
                status_code=404,
                detail="해당 견적서를 찾을 수 없습니다."
            )

        # 2. 견적 제품 목록 조회 (센터명과 함께)
        center_subquery = db.query(set_models.SetCenter.center_name).filter(
            set_models.SetCenter.center_no == purchase_models.OrderShipmentEstimateProduct.center_no,
            set_models.SetCenter.del_yn == 0
        ).scalar_subquery()

        # 포장비닐 사양 정보 조회
        vinyl_spec_subquery = db.query(common_models.ComCode.code_name).filter(
            common_models.ComCode.com_code == purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_cd,
            common_models.ComCode.parent_com_code == 'PACKAGE_VINYL_SPEC_CD',
            common_models.ComCode.del_yn == 0,
            common_models.ComCode.use_yn == 1
        ).scalar_subquery()

        estimate_products = db.query(
            purchase_models.OrderShipmentEstimateProduct,
            center_subquery.label("center_name"),
            vinyl_spec_subquery.label("package_vinyl_spec_name")
        ).filter(
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_no == order_shipment_estimate_no,
            purchase_models.OrderShipmentEstimateProduct.del_yn == 0
        ).order_by(
            purchase_models.OrderShipmentEstimateProduct.fail_yn.asc(),  # 성공한 것부터
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_product_no.asc()
        ).all()

        # 3. 견적 박스 목록 조회 (센터명과 함께)
        box_center_subquery = db.query(set_models.SetCenter.center_name).filter(
            set_models.SetCenter.center_no == purchase_models.OrderShipmentEstimateBox.center_no,
            set_models.SetCenter.del_yn == 0
        ).scalar_subquery()

        # 박스 사양 정보 조회
        box_spec_subquery = db.query(common_models.ComCode.code_name).filter(
            common_models.ComCode.com_code == purchase_models.OrderShipmentEstimateBox.package_box_spec_cd,
            common_models.ComCode.parent_com_code == 'PACKAGE_BOX_SPEC_CD',
            common_models.ComCode.del_yn == 0,
            common_models.ComCode.use_yn == 1
        ).scalar_subquery()

        estimate_boxes = db.query(
            purchase_models.OrderShipmentEstimateBox,
            box_center_subquery.label("center_name"),
            box_spec_subquery.label("package_box_spec_name")
        ).filter(
            purchase_models.OrderShipmentEstimateBox.order_shipment_estimate_no == order_shipment_estimate_no,
            purchase_models.OrderShipmentEstimateBox.del_yn == 0
        ).order_by(
            purchase_models.OrderShipmentEstimateBox.order_shipment_estimate_box_no.asc()
        ).all()

        # 4. 견적 성공 제품 데이터 포맷팅
        product_estimates = []
        for product, center_name, package_vinyl_spec_name in estimate_products:
            if product.fail_yn == 0:  # 성공한 제품만
                product_estimates.append({
                    "order_shipment_mst_no": product.order_shipment_mst_no,
                    "order_shipment_dtl_no": product.order_shipment_dtl_no,
                    "center_no": product.center_no,
                    "center_name": center_name,
                    "sku_name": product.sku_name,
                    "bundle": product.bundle,
                    "quantity": product.purchase_quantity,
                    "sku_id": product.sku_id,
                    "unit_price": float(product.product_unit_price) if product.product_unit_price else 0.0,
                    "product_amount": float(product.product_total_amount) if product.product_total_amount else 0.0,
                    "package_vinyl_spec_cd": product.package_vinyl_spec_cd,
                    "package_vinyl_spec_name": package_vinyl_spec_name,
                    "package_amount": float(product.package_vinyl_spec_total_amount) if product.package_vinyl_spec_total_amount else 0.0,
                    "total_amount": float(product.total_amount) if product.total_amount else 0.0
                })

        # 5. 견적 실패 제품 데이터 포맷팅
        product_estimates_fail = []
        for product, center_name, package_vinyl_spec_name in estimate_products:
            if product.fail_yn == 1:  # 실패한 제품만
                product_estimates_fail.append({
                    "order_shipment_mst_no": product.order_shipment_mst_no,
                    "order_shipment_dtl_no": product.order_shipment_dtl_no,
                    "center_no": product.center_no,
                    "center_name": center_name,
                    "sku_name": product.sku_name,
                    "bundle": product.bundle,
                    "quantity": product.purchase_quantity,
                    "sku_id": product.sku_id,
                    "unit_price": float(product.product_unit_price) if product.product_unit_price else 0.0,
                    "product_amount": float(product.product_total_amount) if product.product_total_amount else 0.0,
                    "package_vinyl_spec_cd": product.package_vinyl_spec_cd,
                    "package_vinyl_spec_name": package_vinyl_spec_name,
                    "package_amount": float(product.package_vinyl_spec_total_amount) if product.package_vinyl_spec_total_amount else 0.0,
                    "total_amount": float(product.total_amount) if product.total_amount else 0.0,
                    "error_message": product.remark  # 실패 사유
                })

        # 6. 박스 견적 데이터 포맷팅
        box_estimates = []
        for box, center_name, package_box_spec_name in estimate_boxes:
            box_estimates.append({
                "center_no": box.center_no,
                "center_name": center_name,
                "package_box_spec_cd": box.package_box_spec_cd,
                "package_box_spec_name": package_box_spec_name,
                "quantity": box.box_quantity if hasattr(box, 'box_quantity') else 1,
                "unit_price": float(box.package_box_spec_unit_price) if box.package_box_spec_unit_price else 0.0,
                "amount": float(box.total_amount) if box.total_amount else 0.0
            })

        # 7. 총 견적 데이터 포맷팅
        total_estimate = {
            "product_total_amount": float(estimate_mst.product_total_amount) if estimate_mst.product_total_amount else 0.0,
            "vinyl_total_amount": float(estimate_mst.vinyl_total_amount) if estimate_mst.vinyl_total_amount else 0.0,
            "box_total_amount": float(estimate_mst.box_total_amount) if estimate_mst.box_total_amount else 0.0,
            "grand_total_amount": float(estimate_mst.estimate_total_amount) if estimate_mst.estimate_total_amount else 0.0
        }

        # 8. 응답 데이터 구성
        response_data = {
            "product_estimates": product_estimates,
            "product_estimates_fail": product_estimates_fail,
            "box_estimates": box_estimates,
            "total_estimate": total_estimate,
            "estimate_info": {
                "order_shipment_estimate_no": estimate_mst.order_shipment_estimate_no,
                "order_mst_no": estimate_mst.order_mst_no,
                "estimate_id": estimate_mst.estimate_id,
                "estimate_date": estimate_mst.estimate_date,
                "deposit_yn": estimate_mst.deposit_yn,
                "created_at": estimate_mst.created_at.isoformat() if estimate_mst.created_at else None,
                "created_by": estimate_mst.created_by
            }
        }

        return common_response.ResponseBuilder.success(
            data=response_data,
            message=f"견적서 상세 정보를 성공적으로 조회했습니다. (견적서 ID: {estimate_mst.estimate_id})"
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=400,
            detail=f"견적서 상세 조회 중 오류가 발생했습니다: {str(e)}"
        )
def confirm_estimate_deposit(
        order_shipment_estimate_no: Union[str, int],
        request: Request,
        db: Session
) -> common_response.ApiResponse[dict]:
    """견적서 입금확인 처리 - 관련된 모든 쉽먼트의 상태를 PAYMENT_COMPLETED로 변경하고 견적서 입금확인"""
    try:
        # 1. 견적서 존재 확인
        estimate = db.query(purchase_models.OrderShipmentEstimate).filter(
            purchase_models.OrderShipmentEstimate.order_shipment_estimate_no == order_shipment_estimate_no,
            purchase_models.OrderShipmentEstimate.del_yn == 0
        ).first()

        if not estimate:
            raise HTTPException(
                status_code=404,
                detail="해당 견적서를 찾을 수 없습니다."
            )

        # 2. 해당 견적서에 속한 견적 상품들에서 order_shipment_mst_no 조회 (중복 제거)
        shipment_mst_nos = db.query(
            purchase_models.OrderShipmentEstimateProduct.order_shipment_mst_no
        ).filter(
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_no == order_shipment_estimate_no,
            purchase_models.OrderShipmentEstimateProduct.del_yn == 0,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_mst_no.isnot(None)  # NULL 제외
        ).distinct().all()

        # order_shipment_mst_no 리스트 추출
        shipment_mst_no_list = [row[0] for row in shipment_mst_nos]

        if not shipment_mst_no_list:
            raise HTTPException(
                status_code=400,
                detail="견적서에 연결된 쉽먼트를 찾을 수 없습니다."
            )

        # 3. 인증된 사용자 정보 가져오기
        user_no, company_no = get_authenticated_user_no(request)

        # 4. 해당 order_shipment_mst_no들의 상태를 PAYMENT_COMPLETED로 업데이트
        updated_count = db.query(purchase_models.OrderShipmentMst).filter(
            purchase_models.OrderShipmentMst.order_shipment_mst_no.in_(shipment_mst_no_list),
            purchase_models.OrderShipmentMst.del_yn == 0
        ).update(
            {
                "order_shipment_mst_status_cd": "PAYMENT_COMPLETED",
                "updated_by": user_no,
                "updated_at": func.now()
            },
            synchronize_session=False
        )

        # 5. 견적서의 deposit_yn을 1로 업데이트 (입금확인)
        estimate.deposit_yn = 1
        estimate.updated_by = user_no
        estimate.updated_at = func.now()

        # 6. 커밋
        db.commit()
        db.refresh(estimate)  # 업데이트된 견적서 정보 새로고침

        # 7. 업데이트된 쉽먼트 정보 조회 (확인용)
        updated_shipments = db.query(purchase_models.OrderShipmentMst).filter(
            purchase_models.OrderShipmentMst.order_shipment_mst_no.in_(shipment_mst_no_list),
            purchase_models.OrderShipmentMst.del_yn == 0
        ).all()

        # 8. 응답 데이터 구성
        response_data = {
            "order_shipment_estimate_no": order_shipment_estimate_no,
            "estimate_id": estimate.estimate_id,
            "deposit_yn": estimate.deposit_yn,
            "updated_shipment_count": updated_count,
            "updated_shipments": [
                {
                    "order_shipment_mst_no": shipment.order_shipment_mst_no,
                    "center_no": shipment.center_no,
                    "display_center_name": shipment.display_center_name,
                    "order_shipment_mst_status_cd": shipment.order_shipment_mst_status_cd,
                    "edd": shipment.edd
                }
                for shipment in updated_shipments
            ]
        }

        return common_response.ResponseBuilder.success(
            data=response_data,
            message=f"견적서 입금확인이 완료되었습니다. (견적서 ID: {estimate.estimate_id}, 업데이트된 쉽먼트: {updated_count}건)"
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"견적서 입금확인 처리 중 오류가 발생했습니다: {str(e)}"
        )


from urllib.parse import quote


async def download_shipment_dtl_excel(
        order_mst_no: Union[str, int],
        request: Request,
        db: Session
) -> FileResponse:
    """Growth 발주 구매 정보 엑셀 다운로드 - 새 파일 생성"""
    try:
        # 발주서 마스터 존재 확인
        existing_order_mst = db.query(purchase_models.OrderMst).filter(
            purchase_models.OrderMst.order_mst_no == order_mst_no,
            purchase_models.OrderMst.del_yn == 0
        ).first()

        if not existing_order_mst:
            raise HTTPException(
                status_code=404,
                detail="해당 발주서를 찾을 수 없습니다."
            )

        # center_name 서브쿼리
        center_subquery = db.query(set_models.SetCenter.center_name).filter(
            set_models.SetCenter.center_no == purchase_models.OrderShipmentMst.center_no,
            set_models.SetCenter.del_yn == 0
        ).scalar_subquery()

        # 데이터 조회
        query = db.query(
            purchase_models.OrderShipmentMst,
            purchase_models.OrderShipmentDtl,
            purchase_models.OrderShipmentPackingDtl,
            purchase_models.OrderShipmentPackingMst,
            center_subquery.label("center_name")
        ).join(
            purchase_models.OrderShipmentDtl,
            purchase_models.OrderShipmentMst.order_shipment_mst_no == purchase_models.OrderShipmentDtl.order_shipment_mst_no
        ).outerjoin(
            purchase_models.OrderShipmentPackingDtl,
            and_(
                purchase_models.OrderShipmentDtl.order_shipment_dtl_no == purchase_models.OrderShipmentPackingDtl.order_shipment_dtl_no,
                purchase_models.OrderShipmentPackingDtl.del_yn == 0
            )
        ).outerjoin(
            purchase_models.OrderShipmentPackingMst,
            and_(
                purchase_models.OrderShipmentPackingDtl.order_shipment_packing_mst_no == purchase_models.OrderShipmentPackingMst.order_shipment_packing_mst_no,
                purchase_models.OrderShipmentPackingMst.del_yn == 0
            )
        ).filter(
            purchase_models.OrderShipmentMst.order_mst_no == order_mst_no,
            purchase_models.OrderShipmentMst.del_yn == 0,
            purchase_models.OrderShipmentDtl.del_yn == 0
        ).order_by(
            purchase_models.OrderShipmentMst.estimated_yn.desc(),
            purchase_models.OrderShipmentDtl.created_at.desc()
        )

        results = query.all()

        if not results:
            raise HTTPException(
                status_code=404,
                detail="다운로드할 데이터가 없습니다."
            )

        # 새 워크북 생성
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "발주 구매 정보"

        # 스타일 정의
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        cell_alignment = Alignment(horizontal="left", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = [
            "발주번호",
            "물류센터",
            "상태",
            "입고유형",
            "입고예정일",
            "상품번호(SKU ID)",
            "상품바코드",
            "상품이름",
            "확정수량",
            "포장수량",
            "박스명",
            "1688 송장번호",
            "CJ 송장번호",
        ]

        # 공통코드
        shipment_status_com_code_dict = com_code_util.get_com_code_dict_by_parent_code("ORDER_SHIPMENT_MST_STATUS_CD",
                                                                                       db)

        # 헤더 작성
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 데이터 작성
        for row_idx, (mst, dtl, packing_dtl, packing_mst, center_name) in enumerate(results, start=2):
            # 발주번호
            cell = worksheet.cell(row=row_idx, column=1, value=dtl.order_number)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 물류센터
            cell = worksheet.cell(row=row_idx, column=2, value=center_name)
            cell.alignment = center_alignment
            cell.border = thin_border

            # 상태
            com_code = shipment_status_com_code_dict.get(mst.order_shipment_mst_status_cd)
            cell = worksheet.cell(row=row_idx, column=3, value=com_code.code_name)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 입고유형
            cell = worksheet.cell(row=row_idx, column=4, value=dtl.transport_type)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 입고예정일
            cell = worksheet.cell(row=row_idx, column=5, value=mst.edd)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # SKU ID
            cell = worksheet.cell(row=row_idx, column=6, value=dtl.sku_id)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 상품바코드
            cell = worksheet.cell(row=row_idx, column=7, value=dtl.sku_barcode)
            cell.alignment = center_alignment
            cell.border = thin_border

            # 상품이름
            cell = worksheet.cell(row=row_idx, column=8, value=dtl.sku_name)
            cell.alignment = center_alignment
            cell.border = thin_border

            # 확정수량
            cell = worksheet.cell(row=row_idx, column=9, value=dtl.confirmed_quantity)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 포장수량
            packing_qty = packing_dtl.packing_quantity if packing_dtl else None
            cell = worksheet.cell(row=row_idx, column=10, value=packing_qty)
            cell.alignment = center_alignment
            cell.border = thin_border

            # 박스명
            box_name = packing_mst.box_name if packing_mst else None
            cell = worksheet.cell(row=row_idx, column=11, value=box_name)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 1688 운송장번호
            cell = worksheet.cell(row=row_idx, column=12, value=dtl.purchase_tracking_number)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # CJ 운송장번호
            tracking_number = packing_dtl.tracking_number if packing_dtl else None
            cell = worksheet.cell(row=row_idx, column=13, value=tracking_number)
            cell.alignment = cell_alignment
            cell.border = thin_border

        # 열 너비 자동 조정
        column_widths = {
            1: 15, 2: 12, 3: 20, 4: 20, 5: 20, 6: 40,
            7: 12, 8: 12, 9: 20, 10: 50, 11: 30, 12: 12, 13: 12,
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = width

        # 임시 파일 생성
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            temp_path = tmp_file.name

        # 워크북 저장
        workbook.save(temp_path)
        workbook.close()

        # 파일명 생성
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"발주구매정보_{order_mst_no}_{current_time}.xlsx"

        encoded_filename = quote(filename)

        response = FileResponse(
            path=temp_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            background=None
        )

        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"

        return response

    except HTTPException:
        raise
    except Exception as e:
        # 임시 파일 정리
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.unlink(temp_path)

        raise HTTPException(
            status_code=500,
            detail=f"엑셀 다운로드 중 오류가 발생했습니다: {str(e)}"
        )


async def download_shipment_estimate_excel(
        order_mst_no: Union[str, int],
        request: Request,
        db: Session
) -> FileResponse:
    """견적 리스트 엑셀 다운로드"""
    try:
        # 발주서 마스터 존재 확인
        existing_order_mst = db.query(purchase_models.OrderMst).filter(
            purchase_models.OrderMst.order_mst_no == order_mst_no,
            purchase_models.OrderMst.del_yn == 0
        ).first()

        if not existing_order_mst:
            raise HTTPException(
                status_code=404,
                detail="해당 발주서를 찾을 수 없습니다."
            )

        # 데이터 조회
        query = db.query(
            purchase_models.OrderShipmentEstimate
        ).filter(
            purchase_models.OrderShipmentEstimate.order_mst_no == order_mst_no,
            purchase_models.OrderShipmentEstimate.del_yn == 0,
        ).order_by(
            purchase_models.OrderShipmentEstimate.created_at.desc()
        )

        results = query.all()

        if not results:
            raise HTTPException(
                status_code=404,
                detail="다운로드할 데이터가 없습니다."
            )

        # 워크북 생성 및 스타일 정의
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "견적 리스트"

        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_alignment = Alignment(horizontal="left", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ["견적서 번호", "견적일자", "견적총액"]

        # 헤더 작성
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 데이터 작성
        for row_idx, estimate in enumerate(results, start=2):
            worksheet.cell(row=row_idx, column=1, value=estimate.estimate_id).alignment = cell_alignment
            worksheet.cell(row=row_idx, column=1).border = thin_border

            worksheet.cell(row=row_idx, column=2, value=estimate.estimate_date).alignment = center_alignment
            worksheet.cell(row=row_idx, column=2).border = thin_border

            worksheet.cell(row=row_idx, column=3, value=estimate.estimate_total_amount).alignment = cell_alignment
            worksheet.cell(row=row_idx, column=3).border = thin_border

        # 열 너비 설정
        worksheet.column_dimensions['A'].width = 15
        worksheet.column_dimensions['B'].width = 12
        worksheet.column_dimensions['C'].width = 20

        # 임시 파일 생성
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            temp_path = tmp_file.name

        workbook.save(temp_path)
        workbook.close()

        # 파일명 생성
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"견적리스트_{order_mst_no}_{current_time}.xlsx"

        # ✅ 한글 파일명 인코딩
        encoded_filename = quote(filename)

        # ✅ FileResponse 반환
        response = FileResponse(
            path=temp_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            background=None
        )

        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"

        return response

    except HTTPException:
        raise
    except Exception as e:
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.unlink(temp_path)

        raise HTTPException(
            status_code=500,
            detail=f"엑셀 다운로드 중 오류가 발생했습니다: {str(e)}"
        )

async def download_shipment_estimate_product_all_excel(
        order_mst_no: Union[str, int],
        request: Request,
        db: Session
) -> FileResponse:
    """견적 상품 전체 목록 엑셀 다운로드"""
    try:

        # 발주서 마스터 존재 확인
        existing_order_mst = db.query(purchase_models.OrderMst).filter(
            purchase_models.OrderMst.order_mst_no == order_mst_no,
            purchase_models.OrderMst.del_yn == 0
        ).first()

        if not existing_order_mst:
            raise HTTPException(
                status_code=404,
                detail="해당 발주서를 찾을 수 없습니다."
            )

        # center_name 서브쿼리
        center_subquery = db.query(set_models.SetCenter.center_name).filter(
            set_models.SetCenter.center_no == purchase_models.OrderShipmentMst.center_no,
            set_models.SetCenter.del_yn == 0
        ).scalar_subquery()

        # 쉽먼트 상태명 서브쿼리
        shipment_status_subquery = db.query(common_models.ComCode.code_name).filter(
            common_models.ComCode.com_code == purchase_models.OrderShipmentMst.order_shipment_mst_status_cd,
            common_models.ComCode.parent_com_code == 'ORDER_SHIPMENT_MST_STATUS_CD',
            common_models.ComCode.del_yn == 0,
            common_models.ComCode.use_yn == 1
        ).scalar_subquery()

        # 포장비닐 사양명 서브쿼리
        vinyl_spec_subquery = db.query(common_models.ComCode.code_name).filter(
            common_models.ComCode.com_code == purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_cd,
            common_models.ComCode.parent_com_code == 'PACKAGE_VINYL_SPEC_CD',
            common_models.ComCode.del_yn == 0,
            common_models.ComCode.use_yn == 1
        ).scalar_subquery()

        # 데이터 조회 (fetch_shipment_estimate_product_list_all과 동일)
        query = (db.query(
            # EstimateProduct 컬럼
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_product_no,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_no,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_mst_no,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no,
            purchase_models.OrderShipmentEstimateProduct.sku_id,
            purchase_models.OrderShipmentEstimateProduct.sku_name,
            purchase_models.OrderShipmentEstimateProduct.purchase_quantity,
            purchase_models.OrderShipmentEstimateProduct.product_unit_price,
            purchase_models.OrderShipmentEstimateProduct.product_total_amount.label("product_product_total_amount"),
            purchase_models.OrderShipmentEstimateProduct.package_vinyl_spec_total_amount,
            purchase_models.OrderShipmentEstimateProduct.total_amount.label("product_total_amount"),
            purchase_models.OrderShipmentEstimateProduct.remark,

            # Estimate 컬럼
            purchase_models.OrderShipmentEstimate.estimate_id,

            # ShipmentMst 컬럼
            purchase_models.OrderShipmentMst.edd,
            purchase_models.OrderShipmentMst.order_shipment_mst_status_cd,
            center_subquery.label("center_name"),
            shipment_status_subquery.label("order_shipment_mst_status_name"),

            # ShipmentDtl 컬럼
            purchase_models.OrderShipmentDtl.order_number,
            purchase_models.OrderShipmentDtl.sku_barcode,
            purchase_models.OrderShipmentDtl.confirmed_quantity.label("dtl_confirmed_quantity"),
            purchase_models.OrderShipmentDtl.transport_type,
            purchase_models.OrderShipmentDtl.purchase_tracking_number,
            purchase_models.OrderShipmentDtl.purchase_order_number,

            # PackingDtl 컬럼
            purchase_models.OrderShipmentPackingDtl.packing_quantity,
            purchase_models.OrderShipmentPackingDtl.box_name,
            purchase_models.OrderShipmentPackingDtl.tracking_number
        ).join(
            purchase_models.OrderShipmentEstimate,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_estimate_no == purchase_models.OrderShipmentEstimate.order_shipment_estimate_no
        ).join(
            purchase_models.OrderShipmentMst,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_mst_no == purchase_models.OrderShipmentMst.order_shipment_mst_no
        ).outerjoin(
            purchase_models.OrderShipmentDtl,
            and_(
                purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no == purchase_models.OrderShipmentDtl.order_shipment_dtl_no,
                purchase_models.OrderShipmentDtl.del_yn == 0
            )
        ).outerjoin(
            purchase_models.OrderShipmentPackingDtl,
            and_(
                purchase_models.OrderShipmentDtl.order_shipment_dtl_no == purchase_models.OrderShipmentPackingDtl.order_shipment_dtl_no,
                purchase_models.OrderShipmentPackingDtl.del_yn == 0
            )
        ).filter(
            purchase_models.OrderShipmentEstimate.order_mst_no == order_mst_no,
            purchase_models.OrderShipmentEstimateProduct.del_yn == 0,
            purchase_models.OrderShipmentEstimate.del_yn == 0,
            purchase_models.OrderShipmentMst.del_yn == 0
        ).order_by(
            purchase_models.OrderShipmentMst.estimated_yn.desc(),
            purchase_models.OrderShipmentEstimateProduct.created_at.desc()
        ))

        results = query.all()

        if not results:
            raise HTTPException(
                status_code=404,
                detail="다운로드할 데이터가 없습니다."
            )

        # 새 워크북 생성
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "견적 상품 목록"

        # 스타일 정의
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        cell_alignment = Alignment(horizontal="left", vertical="center")
        center_alignment = Alignment(horizontal="center", vertical="center")
        right_alignment = Alignment(horizontal="right", vertical="center")

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # 헤더 정의
        headers = [
            "견적서 번호",
            "구매번호",
            "발주번호",
            "물류센터",
            "상태",
            "입고유형",
            "입고예정일",
            "상품번호(SKU ID)",
            "상품바코드",
            "상품이름",
            "확정수량",
            "포장수량",
            "박스명",
            "1688 운송장번호",
            "CJ 운송장번호",
            "비고",
            "단가",
            "제품금액",
            "포장금액",
            "총금액"
        ]

        # 헤더 작성
        for col_idx, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 데이터 작성
        for row_idx, row in enumerate(results, start=2):
            # 견적서 번호
            cell = worksheet.cell(row=row_idx, column=1, value=row.estimate_id)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 구매번호
            cell = worksheet.cell(row=row_idx, column=2, value=row.purchase_order_number)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 발주번호
            cell = worksheet.cell(row=row_idx, column=3, value=row.order_number)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 물류센터
            cell = worksheet.cell(row=row_idx, column=4, value=row.center_name)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 상태
            cell = worksheet.cell(row=row_idx, column=5, value=row.order_shipment_mst_status_name)
            cell.alignment = center_alignment
            cell.border = thin_border

            # 입고유형
            cell = worksheet.cell(row=row_idx, column=6, value=row.transport_type)
            cell.alignment = center_alignment
            cell.border = thin_border

            # 입고예정일
            cell = worksheet.cell(row=row_idx, column=7, value=row.edd)
            cell.alignment = center_alignment
            cell.border = thin_border

            # 상품번호(SKU ID)
            cell = worksheet.cell(row=row_idx, column=8, value=row.sku_id)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 상품바코드
            cell = worksheet.cell(row=row_idx, column=9, value=row.sku_barcode)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 상품이름
            cell = worksheet.cell(row=row_idx, column=10, value=row.sku_name)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 확정수량
            cell = worksheet.cell(row=row_idx, column=11, value=row.dtl_confirmed_quantity)
            cell.alignment = center_alignment
            cell.border = thin_border

            # 포장수량
            cell = worksheet.cell(row=row_idx, column=12, value=row.packing_quantity)
            cell.alignment = center_alignment
            cell.border = thin_border

            # 박스명
            cell = worksheet.cell(row=row_idx, column=13, value=row.box_name)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 송장번호 (1688 운송장번호)
            cell = worksheet.cell(row=row_idx, column=14, value=row.purchase_tracking_number)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.fill = yellow_fill

            # 송장번호 (cj 운송장번호)
            cell = worksheet.cell(row=row_idx, column=15, value=row.tracking_number)
            cell.alignment = cell_alignment
            cell.border = thin_border


            # 비고
            cell = worksheet.cell(row=row_idx, column=16, value=row.remark)
            cell.alignment = cell_alignment
            cell.border = thin_border

            # 단가
            unit_price = float(row.product_unit_price) if row.product_unit_price else 0.0
            cell = worksheet.cell(row=row_idx, column=17, value=unit_price)
            cell.alignment = right_alignment
            cell.border = thin_border
            cell.number_format = '#,##0'

            # 제품금액
            product_amount = float(row.product_product_total_amount) if row.product_product_total_amount else 0.0
            cell = worksheet.cell(row=row_idx, column=18, value=product_amount)
            cell.alignment = right_alignment
            cell.border = thin_border
            cell.number_format = '#,##0'

            # 포장금액
            package_amount = float(row.package_vinyl_spec_total_amount) if row.package_vinyl_spec_total_amount else 0.0
            cell = worksheet.cell(row=row_idx, column=19, value=package_amount)
            cell.alignment = right_alignment
            cell.border = thin_border
            cell.number_format = '#,##0'

            # 총금액
            total_amount = float(row.product_total_amount) if row.product_total_amount else 0.0
            cell = worksheet.cell(row=row_idx, column=20, value=total_amount)
            cell.alignment = right_alignment
            cell.border = thin_border
            cell.number_format = '#,##0'

        # 열 너비 설정
        column_widths = {
            1: 20,  # 견적서 번호
            2: 20,  # 발주번호
            3: 15,  # 물류센터
            4: 15,  # 상태
            5: 12,  # 입고유형
            6: 12,  # 입고예정일
            7: 20,  # 상품번호(SKU ID)
            8: 20,  # 상품바코드
            9: 40,  # 상품이름
            10: 12,  # 확정수량
            11: 12,  # 포장수량
            12: 25,  # 박스명
            13: 20,  # 송장번호
            14: 30,  # 비고
            15: 12,  # 단가
            16: 12,  # 제품금액
            17: 12,  # 포장금액
            18: 12  # 총금액
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[worksheet.cell(row=1, column=col).column_letter].width = width

        # 임시 파일 생성
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
            temp_path = tmp_file.name

        # 워크북 저장
        workbook.save(temp_path)
        workbook.close()

        # 파일명 생성
        current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"견적상품목록_{order_mst_no}_{current_time}.xlsx"

        # ✅ 한글 파일명 인코딩
        encoded_filename = quote(filename)

        # ✅ FileResponse 반환
        response = FileResponse(
            path=temp_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            background=None
        )

        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{encoded_filename}"

        return response
    except HTTPException:
        raise
    except Exception as e:
        # 임시 파일 정리
        if 'temp_path' in locals() and os.path.exists(temp_path):
            os.unlink(temp_path)

        raise HTTPException(
            status_code=500,
            detail=f"엑셀 다운로드 중 오류가 발생했습니다: {str(e)}"
        )


from openpyxl import load_workbook
from fastapi import UploadFile
import io


async def upload_1688_tracking_number(
        order_mst_no: Union[str, int],
        file: UploadFile,
        request: Request,
        db: Session
) -> common_response.ApiResponse[dict]:
    """1688 송장번호 엑셀 업로드 및 업데이트"""
    try:
        # 사용자 인증
        user_no, company_no = get_authenticated_user_no(request)

        # 발주서 마스터 존재 확인
        existing_order_mst = db.query(purchase_models.OrderMst).filter(
            purchase_models.OrderMst.order_mst_no == order_mst_no,
            purchase_models.OrderMst.del_yn == 0
        ).first()

        if not existing_order_mst:
            raise HTTPException(
                status_code=404,
                detail="해당 발주서를 찾을 수 없습니다."
            )

        # 파일 확장자 확인
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(
                status_code=400,
                detail="엑셀 파일(.xlsx, .xls)만 업로드 가능합니다."
            )

        # 파일 내용 읽기
        contents = await file.read()

        # openpyxl로 엑셀 파일 로드
        workbook = load_workbook(io.BytesIO(contents))
        worksheet = workbook.active

        # 업데이트 결과 저장
        update_count = 0
        error_count = 0
        error_details = []

        # 헤더 행 스킵하고 데이터 행부터 읽기 (2행부터)
        for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            try:
                # 컬럼 매핑
                estimate_id = row[0]  # 견적서 번호 (A열)
                purchase_order_number = row[1]  # 구매번호 (B열)
                order_number = row[2]  # 발주번호 (C열)
                center_name = row[3]  # 물류센터 (D열)
                status = row[4]  # 상태 (E열)
                transport_type = row[5]  # 입고유형 (F열)
                edd = row[6]  # 입고예정일 (G열)
                sku_id = row[7]  # 상품번호(SKU ID) (H열)
                sku_barcode = row[8]  # 상품바코드 (I열)
                sku_name = row[9]  # 상품이름 (J열)
                confirmed_quantity = row[10]  # 확정수량 (K열)
                packing_quantity = row[11]  # 포장수량 (L열)
                box_name = row[12]  # 박스명 (M열)
                purchase_tracking_number = row[13]  # 1688 송장번호 (N열)
                tracking_number = row[14]  # CJ 송장번호 (O열)
                remark = row[15]  # 비고 (P열)

                # 필수 필드 체크 (SKU ID와 발주번호는 필수)
                if not sku_id or not order_number:
                    error_details.append({
                        "row": row_idx,
                        "error": "SKU ID 또는 발주번호가 없습니다.",
                        "sku_id": sku_id,
                        "order_number": order_number
                    })
                    error_count += 1
                    continue

                # 송장번호가 없으면 스킵
                if not purchase_tracking_number or str(purchase_tracking_number).strip() == "":
                    continue

                # OrderShipmentDtl에서 해당 레코드 찾기
                # order_mst_no -> OrderShipmentMst -> OrderShipmentDtl
                shipment_dtl = db.query(purchase_models.OrderShipmentDtl).join(
                    purchase_models.OrderShipmentMst,
                    purchase_models.OrderShipmentDtl.order_shipment_mst_no == purchase_models.OrderShipmentMst.order_shipment_mst_no
                ).filter(
                    purchase_models.OrderShipmentMst.order_mst_no == order_mst_no,
                    purchase_models.OrderShipmentDtl.sku_id == sku_id,
                    purchase_models.OrderShipmentDtl.order_number == order_number,
                    purchase_models.OrderShipmentDtl.del_yn == 0,
                    purchase_models.OrderShipmentMst.del_yn == 0
                ).first()

                if not shipment_dtl:
                    error_details.append({
                        "row": row_idx,
                        "error": "해당 SKU ID와 발주번호로 데이터를 찾을 수 없습니다.",
                        "sku_id": sku_id,
                        "order_number": order_number
                    })
                    error_count += 1
                    continue

                # purchase_tracking_number 업데이트
                shipment_dtl.purchase_tracking_number = str(purchase_tracking_number).strip()
                shipment_dtl.updated_by = user_no
                shipment_dtl.updated_at = func.now()

                update_count += 1

            except Exception as e:
                error_details.append({
                    "row": row_idx,
                    "error": f"처리 중 오류: {str(e)}",
                    "sku_id": sku_id if 'sku_id' in locals() else None,
                    "order_number": purchase_tracking_number if 'order_number' in locals() else None
                })
                error_count += 1
                continue

        # 커밋
        db.commit()

        # 응답 데이터 구성
        response_data = {
            "order_mst_no": order_mst_no,
            "total_rows": worksheet.max_row - 1,  # 헤더 제외
            "update_count": update_count,
            "error_count": error_count,
            "error_details": error_details if error_details else None
        }

        if error_count > 0:
            message = f"1688 송장번호 업로드가 부분적으로 완료되었습니다. (성공: {update_count}건, 실패: {error_count}건)"
        else:
            message = f"1688 송장번호 업로드가 완료되었습니다. (총 {update_count}건 업데이트)"

        return common_response.ResponseBuilder.success(
            data=response_data,
            message=message
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"1688 송장번호 업로드 중 오류가 발생했습니다: {str(e)}"
        )


async def issue_cj_tracking_number(
        Issue_tracking_number_request: purchase_schemas.IssueCjTackingNumberRequest,
        request: Request,
        db: Session
) -> common_response.ApiResponse[dict]:
    """CJ 운송장 번호 발급 및 업데이트"""
    try:
        # 사용자 인증
        user_no, company_no = get_authenticated_user_no(request)

        order_shipment_packing_mst_nos = Issue_tracking_number_request.order_shipment_packing_mst_nos
        success_count = 0
        error_count = 0
        error_details = []
        issued_tracking_numbers = []

        for packing_mst_no in order_shipment_packing_mst_nos:
            try:
                # 1. PackingMst 존재 확인
                packing_mst = db.query(purchase_models.OrderShipmentPackingMst).filter(
                    purchase_models.OrderShipmentPackingMst.order_shipment_packing_mst_no == packing_mst_no,
                    purchase_models.OrderShipmentPackingMst.del_yn == 0
                ).first()

                if not packing_mst:
                    error_details.append({
                        "order_shipment_packing_mst_no": packing_mst_no,
                        "box_name": None,
                        "error": "해당 포장 박스를 찾을 수 없습니다."
                    })
                    error_count += 1
                    continue

                # 2. 이미 운송장이 발급된 경우 스킵
                if packing_mst.tracking_number:
                    error_details.append({
                        "order_shipment_packing_mst_no": packing_mst_no,
                        "box_name": packing_mst.box_name,
                        "error": f"이미 운송장이 발급되었습니다. (운송장번호: {packing_mst.tracking_number})"
                    })
                    error_count += 1
                    continue

                # 3. CJ API 호출하여 운송장 번호 발급
                # CJ API 파라미터 구성
                cj_params = {
                    "box_name": packing_mst.box_name,
                    "package_box_spec_cd": packing_mst.package_box_spec_cd,
                    "company_no": packing_mst.company_no,
                }

                # CJ 물류 API 호출
                cj_response = request_cj_logistics_api(
                    db=db,
                    process="/tracking/issue",  # 실제 CJ API 엔드포인트로 수정 필요
                    params=cj_params
                )

                # 4. API 응답 검증 및 운송장 번호 추출
                tracking_number = None

                # ✅ RESULT_CD 체크
                if not cj_response or cj_response.get("RESULT_CD") != "S":
                    error_message = cj_response.get("RESULT_DETAIL", "알 수 없는 오류") if cj_response else "API 응답 없음"
                    error_details.append({
                        "order_shipment_packing_mst_no": packing_mst_no,
                        "box_name": packing_mst.box_name,
                        "error": f"CJ API 호출 실패: {error_message}",
                        "cj_response": cj_response
                    })
                    error_count += 1
                    continue

                # ✅ INVC_NO 추출
                if "DATA" in cj_response and cj_response["DATA"]:
                    tracking_number = cj_response["DATA"].get("INVC_NO")

                if not tracking_number:
                    error_details.append({
                        "order_shipment_packing_mst_no": packing_mst_no,
                        "box_name": packing_mst.box_name,
                        "error": "CJ API에서 운송장 번호(INVC_NO)를 받지 못했습니다.",
                        "cj_response": cj_response
                    })
                    error_count += 1
                    continue

                # 5. PackingMst 업데이트
                packing_mst.tracking_number = tracking_number
                packing_mst.updated_by = user_no
                packing_mst.updated_at = func.now()

                # 6. 해당 PackingMst에 속한 PackingDtl 중 fail_yn이 0인 것만 업데이트
                # ✅ 1단계: fail_yn이 0인 order_shipment_dtl_no를 서브쿼리로 추출
                valid_dtl_nos_subquery = db.query(
                    purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no
                ).filter(
                    purchase_models.OrderShipmentEstimateProduct.fail_yn == 0,
                    purchase_models.OrderShipmentEstimateProduct.del_yn == 0
                ).subquery()

                # ✅ 2단계: 서브쿼리 결과를 사용하여 업데이트 (join 없음)
                updated_dtl_count = db.query(purchase_models.OrderShipmentPackingDtl).filter(
                    purchase_models.OrderShipmentPackingDtl.order_shipment_packing_mst_no == packing_mst_no,
                    purchase_models.OrderShipmentPackingDtl.del_yn == 0,
                    purchase_models.OrderShipmentPackingDtl.order_shipment_dtl_no.in_(valid_dtl_nos_subquery)  # 서브쿼리 사용
                ).update(
                    {
                        "tracking_number": tracking_number,
                        "updated_by": user_no,
                        "updated_at": func.now()
                    },
                    synchronize_session=False
                )

                success_count += 1
                issued_tracking_numbers.append({
                    "order_shipment_packing_mst_no": packing_mst_no,
                    "box_name": packing_mst.box_name,
                    "tracking_number": tracking_number,
                    "updated_dtl_count": updated_dtl_count
                })

            except Exception as e:
                error_details.append({
                    "order_shipment_packing_mst_no": packing_mst_no,
                    "box_name": packing_mst.box_name if 'packing_mst' in locals() else None,
                    "error": f"처리 중 오류: {str(e)}"
                })
                error_count += 1
                continue

        # 커밋 (성공 건이 있을 때만)
        if success_count > 0:
            db.commit()
        else:
            db.rollback()

        # 응답 데이터 구성
        response_data = {
            "total_count": len(order_shipment_packing_mst_nos),
            "success_count": success_count,
            "error_count": error_count,
            "issued_tracking_numbers": issued_tracking_numbers,
            "error_details": error_details if error_details else None
        }

        if error_count > 0 and success_count > 0:
            message = f"CJ 운송장 발급이 부분적으로 완료되었습니다. (성공: {success_count}건, 실패: {error_count}건)"
        elif error_count > 0:
            message = f"CJ 운송장 발급에 실패했습니다. (실패: {error_count}건)"
        else:
            message = f"CJ 운송장 발급이 완료되었습니다. (총 {success_count}건)"

        return common_response.ResponseBuilder.success(
            data=response_data,
            message=message
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"CJ 운송장 발급 중 오류가 발생했습니다: {str(e)}"
        )


async def create_1688_order(
        create_order_request: purchase_schemas.Create1688OrderRequest,
        request: Request,
        db: Session
) -> common_response.ApiResponse[dict]:
    """1688 실제 주문 생성 (판매자별로 분리)"""
    try:
        user_no, company_no = get_authenticated_user_no(request)
        order_shipment_dtl_nos = create_order_request.order_shipment_dtl_nos
        message = create_order_request.message

        if not order_shipment_dtl_nos:
            raise HTTPException(
                status_code=400,
                detail="쉽먼트 DTL 번호가 필요합니다."
            )

        # 1. 견적 상품 정보 조회
        estimate_products = db.query(
            purchase_models.OrderShipmentEstimateProduct,
            purchase_models.OrderShipmentDtl,
            set_models.SetSku
        ).join(
            purchase_models.OrderShipmentDtl,
            purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no ==
            purchase_models.OrderShipmentDtl.order_shipment_dtl_no
        ).join(
            set_models.SetSku,
            purchase_models.OrderShipmentDtl.sku_id == set_models.SetSku.sku_id
        ).filter(
            purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no.in_(order_shipment_dtl_nos),
            purchase_models.OrderShipmentEstimateProduct.fail_yn == 0,
            purchase_models.OrderShipmentEstimateProduct.del_yn == 0,
            purchase_models.OrderShipmentDtl.del_yn == 0,
            purchase_models.OrderShipmentDtl.company_no == company_no,
            set_models.SetSku.del_yn == 0,
            set_models.SetSku.company_no == company_no
        ).all()

        if not estimate_products:
            raise HTTPException(
                status_code=400,
                detail="주문 가능한 견적 상품이 없습니다."
            )

        # 2. ✅ openUid별로 그룹화
        grouped_by_seller = defaultdict(lambda: {
            "cargo_map": defaultdict(int),
            "dtl_nos": []
        })

        for estimate_product, shipment_dtl, set_sku in estimate_products:
            dtl_no = estimate_product.order_shipment_dtl_no

            # SET_SKU 정보 우선 사용
            offer_id = alibaba_1688_util.extract_offer_id_from_link(
                set_sku.link or shipment_dtl.link
            )
            spec_id = set_sku.linked_spec_id or shipment_dtl.linked_spec_id
            open_uid = set_sku.linked_open_uid or shipment_dtl.linked_open_uid
            quantity = estimate_product.purchase_quantity

            if not offer_id or not spec_id or not open_uid:
                raise HTTPException(
                    status_code=400,
                    detail=f"1688 연동 정보가 없는 상품입니다. (SKU ID: {shipment_dtl.sku_id})"
                )

            # ✅ openUid별로 분류
            seller_data = grouped_by_seller[open_uid]
            cargo_key = (offer_id, spec_id)
            seller_data["cargo_map"][cargo_key] += quantity
            if dtl_no not in seller_data["dtl_nos"]:
                seller_data["dtl_nos"].append(dtl_no)

        # 3. ✅ 판매자별로 주문 생성
        created_orders = []
        total_success = 0
        total_error = 0
        error_details = []

        for open_uid, seller_data in grouped_by_seller.items():
            try:
                # cargo_list 생성 (같은 판매자 상품만)
                cargo_list = []
                for (offer_id, spec_id), total_quantity in seller_data["cargo_map"].items():
                    cargo_list.append(
                        common_schemas.AlibabaFastCreateOrderCargo(
                            offerId=offer_id,
                            specId=spec_id,
                            quantity=total_quantity
                        )
                    )

                # 외부 주문 ID 생성
                dtl_nos = seller_data["dtl_nos"]
                sorted_dtl_nos = sorted(dtl_nos)
                out_order_id = f"DTL_{open_uid[:8]}_{sorted_dtl_nos[0]}"

                # 1688 API 호출
                api_request = common_schemas.AlibabaFastCreateOrderRequest(
                    cargoList=cargo_list,
                    flow="general",
                    message=message,
                    tradeType="creditBuy",
                    outOrderId=out_order_id
                )

                api_result = await alibaba_1688_util.create_order_1688(api_request)

                # API 결과 확인
                if not api_result or not api_result.get("success"):
                    error_message = api_result.get("message", "알 수 없는 오류") if api_result else "API 응답 없음"
                    translated_message = await alibaba_1688_util.translate_chinese_to_korean(error_message)

                    error_details.append({
                        "open_uid": open_uid,
                        "dtl_nos": dtl_nos,
                        "error": f"1688 주문 생성 실패: {translated_message}",
                        "api_response": api_result
                    })
                    total_error += 1
                    continue

                # 주문 ID 추출
                order_id = None
                if "result" in api_result:
                    order_id = api_result["result"].get("orderId")

                if not order_id:
                    error_details.append({
                        "open_uid": open_uid,
                        "dtl_nos": dtl_nos,
                        "error": "1688에서 주문 ID를 받지 못했습니다.",
                        "api_response": api_result
                    })
                    total_error += 1
                    continue

                # DB 업데이트 - 해당 판매자의 상품만
                updated_estimate_count = db.query(purchase_models.OrderShipmentEstimateProduct).filter(
                    purchase_models.OrderShipmentEstimateProduct.order_shipment_dtl_no.in_(dtl_nos),
                    purchase_models.OrderShipmentEstimateProduct.fail_yn == 0,
                    purchase_models.OrderShipmentEstimateProduct.del_yn == 0
                ).update(
                    {
                        "purchase_order_number": order_id,
                        "updated_by": user_no,
                        "updated_at": func.now()
                    },
                    synchronize_session=False
                )

                updated_dtl_count = db.query(purchase_models.OrderShipmentDtl).filter(
                    purchase_models.OrderShipmentDtl.order_shipment_dtl_no.in_(dtl_nos),
                    purchase_models.OrderShipmentDtl.del_yn == 0
                ).update(
                    {
                        "purchase_order_number": order_id,
                        "updated_by": user_no,
                        "updated_at": func.now()
                    },
                    synchronize_session=False
                )

                created_orders.append({
                    "open_uid": open_uid,
                    "order_1688_id": order_id,
                    "dtl_nos": dtl_nos,
                    "total_items": len(cargo_list),
                    "total_quantity": sum(cargo.quantity for cargo in cargo_list),
                    "updated_estimate_count": updated_estimate_count,
                    "updated_dtl_count": updated_dtl_count
                })

                total_success += 1

            except Exception as e:
                error_details.append({
                    "open_uid": open_uid,
                    "dtl_nos": seller_data["dtl_nos"],
                    "error": f"처리 중 오류: {str(e)}"
                })
                total_error += 1
                continue

        # 커밋
        if total_success > 0:
            db.commit()
        else:
            db.rollback()

        # 응답 데이터 구성
        response_data = {
            "total_sellers": len(grouped_by_seller),
            "success_count": total_success,
            "error_count": total_error,
            "created_orders": created_orders,
            "error_details": error_details if error_details else None,
            "total_dtl_count": len(order_shipment_dtl_nos)
        }

        if total_error > 0 and total_success > 0:
            message_text = f"1688 주문이 부분적으로 생성되었습니다. (성공: {total_success}개 판매자, 실패: {total_error}개 판매자)"
        elif total_error > 0:
            message_text = f"1688 주문 생성에 실패했습니다. (실패: {total_error}개 판매자)"
        else:
            message_text = f"1688 주문이 생성되었습니다. (총 {total_success}개 판매자, {len(created_orders)}개 주문)"

        return common_response.ResponseBuilder.success(
            data=response_data,
            message=message_text
        )

    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=400,
            detail=f"1688 주문 생성 중 오류가 발생했습니다: {str(e)}"
        )